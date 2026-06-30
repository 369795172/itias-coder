"""Session persistence: autosave JSON, export Excel/TXT, resume from Excel."""
from __future__ import annotations

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import CodeDef, Profile, Segment, Session

SAVE_FILENAME = ".itias_save.json"


# ── Autosave (JSON sidecar) ───────────────────────────────────────────────────

def autosave_path(segments_folder: str) -> str:
    return os.path.join(segments_folder, SAVE_FILENAME)


def save_session(session: Session) -> None:
    path = autosave_path(session.segments_folder)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(session.to_dict(), f, ensure_ascii=False, indent=2)


def load_session(segments_folder: str) -> Optional[Session]:
    """Load saved session from folder. Returns None if no save file."""
    path = autosave_path(segments_folder)
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        return Session.from_dict(data)
    except Exception:
        return None


def has_save(segments_folder: str) -> bool:
    return os.path.exists(autosave_path(segments_folder))


# ── Export ────────────────────────────────────────────────────────────────────

def export_excel(session: Session, profile: Profile, out_path: str) -> None:
    """Export coding results to Excel (.xlsx), CCIES-compatible structure."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ITIAS编码"

    # Header row
    headers = ["片段序号", "文件名", "编码", "编码描述", "类别", "编码时间"]
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for seg in session.segments:
        code_def = profile.get_code(seg.code_id) if seg.code_id else None
        category_label = ""
        if code_def:
            category_label = profile.category_labels.get(code_def.category, code_def.category)

        row = [
            seg.index,
            seg.filename,
            seg.code_id or "",
            seg.code_label or "",
            category_label,
            seg.coded_at or "",
        ]
        ws.append(row)

    # Apply alternating row colors
    for row_idx in range(2, ws.max_row + 1):
        fill = PatternFill("solid", fgColor="E8F0FE") if row_idx % 2 == 0 else None
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill

    # Column widths
    widths = [8, 30, 6, 20, 16, 24]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Summary sheet
    ws2 = wb.create_sheet("汇总")
    ws2.append(["统计项", "数值"])
    ws2.append(["总片段数", session.total])
    ws2.append(["已编码片段数", session.coded_count])
    ws2.append(["编码完成率", f"{session.coded_count/session.total*100:.1f}%" if session.total else "N/A"])
    ws2.append(["编码框架", profile.name])
    ws2.append(["参考文献", profile.reference])
    ws2.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    ws2.append([])
    ws2.append(["编码", "标签", "出现次数", "占比"])
    code_counts: dict[int, int] = {}
    for seg in session.segments:
        if seg.code_id is not None:
            code_counts[seg.code_id] = code_counts.get(seg.code_id, 0) + 1
    for code_def in profile.codes:
        count = code_counts.get(code_def.id, 0)
        pct = f"{count/session.coded_count*100:.1f}%" if session.coded_count else "0%"
        ws2.append([code_def.id, code_def.label, count, pct])

    wb.save(out_path)


def export_txt(session: Session, out_path: str) -> None:
    """Export code sequence as plain text, one code per line."""
    lines = []
    for seg in session.segments:
        lines.append(str(seg.code_id) if seg.code_id is not None else "")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))


# ── Resume from Excel ─────────────────────────────────────────────────────────

def import_from_excel(excel_path: str, segments: list[Segment]) -> list[Segment]:
    """
    Import codes from a previously exported Excel file into segment list.
    Matches by segment index (column A). Returns updated segment list.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    code_map: dict[int, tuple[Optional[int], Optional[str], Optional[str]]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        seg_index = int(row[0])
        code_id = int(row[2]) if row[2] else None
        code_label = str(row[3]) if row[3] else None
        coded_at = str(row[5]) if row[5] else None
        code_map[seg_index] = (code_id, code_label, coded_at)

    for seg in segments:
        if seg.index in code_map:
            code_id, code_label, coded_at = code_map[seg.index]
            seg.code_id = code_id
            seg.code_label = code_label
            seg.coded_at = coded_at

    return segments
