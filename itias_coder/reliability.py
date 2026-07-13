"""Inter-coder reliability: agreement metrics between two Excel exports."""
from __future__ import annotations

import os
from datetime import datetime
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

from .profile import load_profile

CODE_IDS = list(range(1, 19))
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def load_codes_from_excel(path: str) -> dict[int, Optional[int]]:
    """Parse an exported Excel file; return {segment_index: code_id}."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    codes: dict[int, Optional[int]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        seg_index = int(row[0])
        code_id = int(row[2]) if row[2] else None
        codes[seg_index] = code_id
    return codes


def _category_key(code_id: Optional[int]) -> int:
    """Map code_id to category key for kappa (0 = uncoded)."""
    return code_id if code_id is not None else 0


def _cohens_kappa(pairs: list[tuple[Optional[int], Optional[int]]]) -> Optional[float]:
    """Cohen's Kappa for nominal ratings (uncoded segments use category 0)."""
    if not pairs:
        return None

    categories = sorted({_category_key(a) for a, _ in pairs} | {_category_key(b) for _, b in pairs})
    cat_index = {cat: i for i, cat in enumerate(categories)}
    n = len(pairs)
    table = [[0] * len(categories) for _ in categories]

    for a, b in pairs:
        table[cat_index[_category_key(a)]][cat_index[_category_key(b)]] += 1

    observed = sum(table[i][i] for i in range(len(categories))) / n
    row_margins = [sum(row) for row in table]
    col_margins = [sum(table[r][c] for r in range(len(categories))) for c in range(len(categories))]
    expected = sum(row_margins[i] * col_margins[i] for i in range(len(categories))) / (n * n)

    if expected >= 1.0:
        return 1.0 if observed >= 1.0 else 0.0
    return (observed - expected) / (1.0 - expected)


def compute_agreement(
    codes1: dict[int, Optional[int]],
    codes2: dict[int, Optional[int]],
) -> dict:
    """Compute percentage agreement, Cohen's Kappa, and 18x18 confusion matrix."""
    all_indices = sorted(set(codes1) | set(codes2))
    pairs = [(codes1.get(i), codes2.get(i)) for i in all_indices]

    agreed = sum(1 for a, b in pairs if a == b)
    total = len(pairs)
    overall_pct = (agreed / total * 100.0) if total else 0.0

    confusion: dict[int, dict[int, int]] = {i: {j: 0 for j in CODE_IDS} for i in CODE_IDS}
    for a, b in pairs:
        if a in CODE_IDS and b in CODE_IDS:
            confusion[a][b] += 1

    per_code: dict[int, dict] = {}
    for code_id in CODE_IDS:
        c1 = sum(1 for a, _ in pairs if a == code_id)
        c2 = sum(1 for _, b in pairs if b == code_id)
        both = sum(1 for a, b in pairs if a == code_id and b == code_id)
        either = sum(1 for a, b in pairs if a == code_id or b == code_id)
        per_code[code_id] = {
            "coder1_count": c1,
            "coder2_count": c2,
            "both_count": both,
            "either_count": either,
            "agreement_pct": (both / either * 100.0) if either else 0.0,
        }

    segment_rows = [
        {
            "index": idx,
            "coder1": codes1.get(idx),
            "coder2": codes2.get(idx),
            "agree": codes1.get(idx) == codes2.get(idx),
        }
        for idx in all_indices
    ]

    coded_both = sum(1 for a, b in pairs if a is not None and b is not None)
    coded_agreed = sum(1 for a, b in pairs if a is not None and b is not None and a == b)
    coded_pct = (coded_agreed / coded_both * 100.0) if coded_both else 0.0

    return {
        "total_segments": total,
        "agreed_segments": agreed,
        "overall_agreement_pct": overall_pct,
        "coded_segments": coded_both,
        "coded_agreement_pct": coded_pct,
        "cohens_kappa": _cohens_kappa(pairs),
        "confusion_matrix": confusion,
        "per_code": per_code,
        "segments": segment_rows,
    }


def export_report(results: dict, out_path: str) -> None:
    """Write reliability report to Excel (data + summary + per-code sheets)."""
    profile = load_profile("itias_default")
    wb = openpyxl.Workbook()

    # ── Data sheet ────────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "编码对比"
    headers = ["片段序号", "编码者1", "编码者2", "是否一致", "编码1描述", "编码2描述"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, seg in enumerate(results["segments"], 2):
        c1 = seg["coder1"]
        c2 = seg["coder2"]
        d1 = profile.get_code(c1).label if c1 and profile.get_code(c1) else ""
        d2 = profile.get_code(c2).label if c2 and profile.get_code(c2) else ""
        ws.append([
            seg["index"],
            c1 if c1 is not None else "",
            c2 if c2 is not None else "",
            "是" if seg["agree"] else "否",
            d1,
            d2,
        ])
        fill = PatternFill("solid", fgColor="E8F0FE") if row_idx % 2 == 0 else None
        if fill:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    widths = [8, 10, 10, 10, 20, 20]
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws2 = wb.create_sheet("汇总")
    ws2.append(["统计项", "数值"])
    ws2.append(["总片段数", results["total_segments"]])
    ws2.append(["一致片段数", results["agreed_segments"]])
    ws2.append(["百分比一致性（全部）", f"{results['overall_agreement_pct']:.1f}%"])
    ws2.append(["双编码片段数", results["coded_segments"]])
    ws2.append(["百分比一致性（仅双编码）", f"{results['coded_agreement_pct']:.1f}%"])
    kappa = results["cohens_kappa"]
    ws2.append(["Cohen's Kappa", f"{kappa:.4f}" if kappa is not None else "N/A"])
    ws2.append(["编码框架", profile.name])
    ws2.append(["导出时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

    # ── Per-code sheet ────────────────────────────────────────────────────────
    ws3 = wb.create_sheet("分编码一致性")
    per_code_headers = ["编码", "标签", "编码者1次数", "编码者2次数", "双方一致次数", "一致率"]
    for col, h in enumerate(per_code_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for code_def in profile.codes:
        stats = results["per_code"][code_def.id]
        ws3.append([
            code_def.id,
            code_def.label,
            stats["coder1_count"],
            stats["coder2_count"],
            stats["both_count"],
            f"{stats['agreement_pct']:.1f}%",
        ])

    # ── Confusion matrix sheet ────────────────────────────────────────────────
    ws4 = wb.create_sheet("混淆矩阵")
    ws4.cell(row=1, column=1, value="编码者1 \\ 编码者2")
    for col, code_id in enumerate(CODE_IDS, 2):
        cell = ws4.cell(row=1, column=col, value=code_id)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for row_idx, row_code in enumerate(CODE_IDS, 2):
        ws4.cell(row=row_idx, column=1, value=row_code)
        for col_idx, col_code in enumerate(CODE_IDS, 2):
            ws4.cell(
                row=row_idx,
                column=col_idx,
                value=results["confusion_matrix"][row_code][col_code],
            )

    wb.save(out_path)


def main_cli(file1: str, file2: str, out_path: str) -> None:
    """CLI entry: load two exports, compute metrics, write report."""
    codes1 = load_codes_from_excel(file1)
    codes2 = load_codes_from_excel(file2)
    if not codes1:
        raise ValueError(f"编码者1文件无有效数据: {file1}")
    if not codes2:
        raise ValueError(f"编码者2文件无有效数据: {file2}")

    results = compute_agreement(codes1, codes2)
    if not out_path:
        base = os.path.splitext(os.path.basename(file1))[0]
        out_path = os.path.join(os.path.dirname(file1), f"{base}_reliability.xlsx")
    export_report(results, out_path)
